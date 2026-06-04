from __future__ import annotations

import asyncio
import logging

from ai_file_brain.config import AiFileBrainSettings
from ai_file_brain.core.models import ExtractionResult

logger = logging.getLogger(__name__)


class XlsxExtractor:
    async def extract(self, file_path: str) -> ExtractionResult:
        cap = AiFileBrainSettings().max_extracted_chars
        text = await asyncio.to_thread(self._read_sync, file_path, cap)
        return ExtractionResult(text=text, source="native")

    @staticmethod
    def _read_sync(file_path: str, max_chars: int) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError as ex:
            logger.warning("openpyxl not available; cannot read %s: %s", file_path, ex)
            return ""

        # read_only keeps memory flat on big sheets; data_only returns the last
        # cached cell *values* instead of formula strings.
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except FileNotFoundError:
            return ""
        except Exception as ex:
            # Password-protected, malformed, or a legacy .xls misnamed as .xlsx.
            logger.warning("Failed to open XLSX %s: %s", file_path, ex)
            return ""

        parts: list[str] = []
        total = 0  # running char count, to stop flattening a giant sheet into RAM
        truncated = False
        try:
            for ws in wb.worksheets:
                if truncated:
                    break
                rows: list[str] = []
                for row in ws.iter_rows(values_only=True):
                    cells = [str(v) for v in row if v is not None and str(v).strip()]
                    if cells:
                        line = "\t".join(cells)
                        rows.append(line)
                        total += len(line) + 1
                        if max_chars and total >= max_chars:
                            truncated = True
                            break
                if rows:
                    # Title the sheet's block so retrieval keeps per-sheet context.
                    parts.append(f"# {ws.title}")
                    parts.extend(rows)
        finally:
            wb.close()

        if truncated:
            logger.info(
                "Truncating %s at ~%d chars (max_extracted_chars)", file_path, max_chars
            )
        return "\n".join(parts)
